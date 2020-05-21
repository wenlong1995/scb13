#!/usr/bin/env python 
# -*- coding: utf-8 -*- 
# @Time : 2020/5/21 15:02
# @Author : Lemon_Youzi
# @QQ: 1433281925
# Copyright：
'''
接口项目实战
步骤：1.连接文档 2.读取文档数据 3.执行测试用例，并预期结果与实际结果比对 4.写入结果，保存文件
'''
import requests
import openpyxl
session = requests.session()
#读取文档数据的函数
def read(test_case,sheetname):
    wb = openpyxl.load_workbook(test_case)   #加载工作簿
    sheet = wb[sheetname]  #加载表单
    max_row = sheet.max_row  #获取最大行  最大列表示：max_column
    cases = []  #定义一个空列表进行数据存储
    #for循环可以遍历所有的内容
    for i in range(2,max_row+1):   #range函数取头不取尾，所以需要+1
        case = dict(                            #以字典形式生成
        case_id = sheet.cell(row=i,column=1).value,  #
        url = sheet.cell(row=i, column=5).value,   #获取URL
        data = sheet.cell(row=i, column=6).value,  #获取data参数数据
        expected_result = sheet.cell(row=i, column=7).value) #获取预期结果
        cases.append(case)  #添加元素到列表
    return (cases)
#写入结果的函数
def xr(test_case,sheetname,row,column,real_result):
    wb = openpyxl.load_workbook(test_case)  #加载工作簿
    sheet = wb[sheetname]              #加载表单
    cell = sheet.cell(row=row,column=column).value = real_result  #写入数据，重新赋值
    wb.save(test_case)
#发送接口请求的函数
def funtion(url,data):
    response = session.post(url=url,data=data)
    result = response.json()
    return (result)
#用例执行函数
def execute_func(test_case,sheetname):
    cases = read(test_case,sheetname)   #调用读取文档函数
    for case in cases:
        case_id = case.get('case_id')  #获取case_id
        url = case.get('url')         #获取URL
        data = case.get('data')         #获取data，以dict.get()方式
        data = eval(data)             #使用eval（）函数，将字符串转为字典格式
        expected_result = case['expected_result']     #获取预期结果
        print(type('expected_result'))      #判断数据类型
        expected_result = expected_result.replace('null','None')  #使用字符串replace函数替换内容
        expected_result = eval(expected_result)     #字符串转为字典
        real_result = funtion(url,data)     #调用发送请求函数
        real_msg = real_result['msg']   #提取msg的信息
        expecte_msg = expected_result['msg']      #获取msg信息
        print('真实期望结果：{}'.format(real_msg))    #格式化输出：便于阅读
        print('预期测试结果：{}'.format(expecte_msg))
        if real_msg== expecte_msg:     #判断预期结果与执行结果是否一致
            print('第{}条测试用例通过：'.format(case_id))
            final_result = 'passed'
            print('***'*10)
        else:
            print('第{}条测试用例不通过：'.format(case_id))
            final_result = 'falsed'
            print('***' * 10)
        xr(test_case,sheetname,case_id+1,8,final_result)   #调用写入数据函数
execute_func('test_case.xlsx','login')       #调用执行用例函数

